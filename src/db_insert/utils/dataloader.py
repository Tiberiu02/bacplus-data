import json
from openpyxl import load_workbook
import csv


def load_data_file(filename):
    print("Loading data...")
    if filename.endswith(".json"):
        return json.load(open(filename, "r", encoding="utf-8"))
    elif filename.endswith(".xlsx"):
        workbook = load_workbook(filename=filename)
        sheet = workbook.active
        raw_data = [[str(cell.value) for cell in row] for row in sheet.iter_rows()]
        workbook.close()
    elif filename.endswith(".csv"):
        with open(filename, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            raw_data = list(reader)
    else:
        raise ValueError(f"Unsupported file format: {filename}")

    num_rows = len(raw_data)
    num_cols = max(len(row) for row in raw_data)

    print("num_cols:", num_cols)

    start_row = 0
    while start_row < num_rows and raw_data[start_row][2] is None:
        start_row += 1

    header = raw_data[start_row]
    rows = raw_data[start_row + 1 :]

    entries = []
    for row in rows:
        entry = {}
        for i, cell in enumerate(row):
            entry[header[i]] = cell
        entries.append(entry)

    return entries
