from openpyxl import load_workbook
import re
from dotenv import load_dotenv
import os
import psycopg2

load_dotenv()

# Load XLSX file

workbook = load_workbook(filename="data/siiir/2024_2025_retea_scolara.xlsx")
sheet = workbook.active

num_rows = sheet.max_row
num_cols = sheet.max_column

start_row = 1
while start_row <= num_rows and sheet.cell(row=start_row, column=2).value == None:
    start_row += 1

# Read data

header = [sheet.cell(row=start_row, column=i).value for i in range(1, num_cols + 1)]
column_names = [h.replace(".", "").replace(" ", "_").lower() for h in header]

siiir_columns = [i for i, h in enumerate(header) if "siiir" in str(h).lower()]
print(f"SIIIR columns: {siiir_columns}")

data = []
for i in range(start_row + 1, num_rows + 1):
    row = []
    for j in range(1, num_cols + 1):
        cell = sheet.cell(row=i, column=j).value
        if cell == "":
            cell = None
        # dd/mm/yyyy -> yyyy-mm-dd
        if re.match(r"\d{2}/\d{2}/\d{4}", str(cell)):
            cell = re.sub(r"(\d{2})/(\d{2})/(\d{4})", r"\3-\2-\1", str(cell))
        row.append(cell)
    for j in siiir_columns:
        if row[j] is not None and len(row[j]) > 3:
            row[j] = row[j][:3] + "1" + row[j][4:]
    data.append(row)

# Insert data into database

if os.getenv("DATABASE_URL") is None:
    print("Make sure to specify DATABASE_URL in .env file")
    exit(1)

print(f"Connecting to database...")

conn = psycopg2.connect(os.getenv("DATABASE_URL"))
cur = conn.cursor()

cur.execute("SELECT COUNT(*) FROM public.siiir")
cnt = cur.fetchone()[0]

print(f"Deleting {cnt} entries from database...")

cur.execute("DELETE FROM public.siiir")

print(f"Inserting {len(data)} entries into database...")

cur.execute(
    f"INSERT INTO public.siiir({','.join(column_names)}) VALUES "
    + ",".join(
        cur.mogrify(
            "(" + ",".join(["%s"] * len(x)) + ")",
            x,
        ).decode("utf-8")
        for x in data
    )
)

conn.commit()
conn.close()
