import argparse
import json
import math
from openpyxl import load_workbook
import re
from dotenv import load_dotenv
import os
import psycopg2
import unidecode
from tqdm import tqdm
import csv

load_dotenv()


def validate_schema(schema):
    required_keys = [
        "cod_candidat",
        "specializare",
        "lr_init",
        "lr_cont",
        "limba_materna",
        "lm_init",
        "lm_cont",
        "disciplina_obligatorie",
        "do_init",
        "do_cont",
        "disciplina_alegere",
        "da_init",
        "da_cont",
        "limba_moderna",
        "medie",
        "rezultat",
    ]
    possible_keys = required_keys + [
        "promotie",
        "promotie_anterioara",
        "unitate_siiir",
        "unitate_nume",
        "unitate_cod_judet",
        "sex",
        "clasa",
    ]

    # Check required keys
    for key in required_keys:
        if key not in schema:
            raise ValueError(f"Missing key '{key}' in schema")

    # Check for unknown keys
    for key in schema:
        if key not in possible_keys:
            raise ValueError(f"Unknown key '{key}' in schema")

    # Check promotie
    if not ("promotie" in schema or "promotie_anterioara" in schema):
        raise ValueError("Missing 'promotie' or 'promotie_anterioara' key in schema")
    if "promotie" in schema and "promotie_anterioara" in schema:
        raise ValueError("Both 'promotie' and 'promotie_anterioara' keys in schema")

    # Check unitate
    if not ("unitate_siiir" in schema or "unitate_nume" in schema):
        raise ValueError("Missing 'unitate_siiir' or 'unitate_nume' key in schema")
    if "unitate_siiir" in schema and "unitate_nume" in schema:
        raise ValueError("Both 'unitate_siiir' and 'unitate_nume' keys in schema")
    if "unitate_siiir" in schema and "unitate_cod_judet" in schema:
        raise ValueError("Both 'unitate_siiir' and 'unitate_cod_judet' keys in schema")
    if "unitate_nume" in schema and not "unitate_cod_judet" in schema:
        raise ValueError(
            "Key 'unitate_cod_judet' required when 'unitate_nume' is present in schema"
        )


def parse_header(header, schema):
    columns = {}
    for key in schema:
        col_name = schema[key]
        if col_name not in header:
            raise ValueError(
                f"Missing column '{col_name}' in header (schema key '{key}')"
            )
        columns[key] = header.index(col_name)
    return columns


parser = argparse.ArgumentParser()
parser.add_argument("year", type=int)
parser.add_argument("data_file", type=str)
parser.add_argument("schema_file", type=str)
args = parser.parse_args()

print("Loading schema...")
schema = json.load(open(args.schema_file, "r", encoding="utf-8"))
validate_schema(schema)

# Load XLSX file

print("Loading data...")
if args.data_file.endswith(".xlsx"):
    workbook = load_workbook(filename=args.data_file)
    sheet = workbook.active
    raw_data = [[str(cell.value) for cell in row] for row in sheet.iter_rows()]
    workbook.close()
elif args.data_file.endswith(".csv"):
    with open(args.data_file, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        raw_data = list(reader)
else:
    raise ValueError("Invalid data file format")

num_rows = len(raw_data)
num_cols = max(len(row) for row in raw_data)

print("num_cols:", num_cols)

start_row = 0
while start_row < num_rows and raw_data[start_row][2] is None:
    start_row += 1

# Read data

header = raw_data[start_row]
columns = parse_header(header, schema)

print(header)
print(columns)

data = []
data_cols = [
    "an",
    "cod_candidat",
    "sex",
    "specializare",
    "promotie_anterioara",
    "unitate_siiir",
    "unitate_nume",
    "unitate_cod_judet",
    "clasa",
    "lr_init",
    "lr_cont",
    "lr_final",
    "limba_materna",
    "lm_init",
    "lm_cont",
    "lm_final",
    "disciplina_obligatorie",
    "do_init",
    "do_cont",
    "do_final",
    "disciplina_alegere",
    "da_init",
    "da_cont",
    "da_final",
    "limba_moderna",
    "medie",
    "rezultat",
]


def parse_grade(grade):
    try:
        g = float(grade)
        if g >= 1 and g <= 10:
            return g
    except:
        pass
    return None


def parse_row(row):
    an = args.year
    cod_candidat = row[columns["cod_candidat"]]
    specializare = row[columns["specializare"]]

    clasa = row[columns["clasa"]] if "clasa" in columns else None

    sex = row[columns["sex"]] if "sex" in columns else None
    if sex is not None:
        sex = sex.lower()
        if sex == "masculin":
            sex = "m"
        elif sex == "feminin":
            sex = "f"
        if sex not in ["m", "f"]:
            raise ValueError(f"Invalid sex '{sex}' for candidate {cod_candidat}")

    promotie_anterioara = None
    if "promotie" in columns:
        promotie = row[columns["promotie"]]
        promotie_anterioara = promotie != f"{args.year - 1}-{args.year}"
    else:
        promotie_anterioara = row[columns["promotie_anterioara"]].upper() == "DA"

    unitate_siiir = (
        row[columns["unitate_siiir"]] if "unitate_siiir" in columns else None
    )
    unitate_nume = row[columns["unitate_nume"]] if "unitate_nume" in columns else None
    unitate_cod_judet = (
        row[columns["unitate_cod_judet"]] if "unitate_cod_judet" in columns else None
    )

    lr_init = parse_grade(row[columns["lr_init"]])
    lr_cont = parse_grade(row[columns["lr_cont"]])
    lr_final = lr_cont if lr_cont is not None else lr_init

    limba_materna = row[columns["limba_materna"]]
    lm_init = parse_grade(row[columns["lm_init"]])
    lm_cont = parse_grade(row[columns["lm_cont"]])
    lm_final = lm_cont if lm_cont is not None else lm_init

    disciplina_obligatorie = row[columns["disciplina_obligatorie"]]
    do_init = parse_grade(row[columns["do_init"]])
    do_cont = parse_grade(row[columns["do_cont"]])
    do_final = do_cont if do_cont is not None else do_init

    disciplina_alegere = row[columns["disciplina_alegere"]]
    da_init = parse_grade(row[columns["da_init"]])
    da_cont = parse_grade(row[columns["da_cont"]])
    da_final = da_cont if da_cont is not None else da_init

    limba_moderna = row[columns["limba_moderna"]]

    medie = parse_grade(row[columns["medie"]])
    if (
        lr_final is not None
        and do_final is not None
        and da_final is not None
        and (lm_final is not None or limba_materna is None)
    ):
        sum_grades = sum([lr_final, do_final, da_final, lm_final or 0])
        num_grades = 4 if lm_final is not None else 3

        my_medie = sum_grades / num_grades
        my_medie = (my_medie + 1e-5) // 0.01 / 100

        if medie is not None and medie != my_medie:
            raise ValueError(
                f"Computed average {my_medie} does not match provided average {medie} for candidate {cod_candidat}"
            )
        elif medie is None:
            medie = my_medie

    rezultat = row[columns["rezultat"]].lower()
    if rezultat == "reusit":
        rezultat = "promovat"
    elif rezultat == "respins":
        rezultat = "nepromovat"
    elif rezultat == "neprezentat":
        rezultat = "absent"
    elif rezultat == "eliminat din examen":
        rezultat = "eliminat"
    if rezultat not in ["promovat", "nepromovat", "absent", "eliminat"]:
        raise ValueError(f"Invalid result '{rezultat}' for candidate {cod_candidat}")

    return [
        an,
        cod_candidat,
        sex,
        specializare,
        promotie_anterioara,
        unitate_siiir,
        unitate_nume,
        unitate_cod_judet,
        clasa,
        lr_init,
        lr_cont,
        lr_final,
        limba_materna,
        lm_init,
        lm_cont,
        lm_final,
        disciplina_obligatorie,
        do_init,
        do_cont,
        do_final,
        disciplina_alegere,
        da_init,
        da_cont,
        da_final,
        limba_moderna,
        medie,
        rezultat,
    ]


print("Parsing data...")
for i in range(start_row + 1, num_rows):
    row = raw_data[i]
    data.append(parse_row(row))

# Insert data into database

if os.getenv("DATABASE_URL") is None:
    print("Make sure to specify DATABASE_URL in .env file")
    exit(1)

print(f"Connecting to database...")

conn = psycopg2.connect(os.getenv("DATABASE_URL"))
cur = conn.cursor()

cur.execute(f"SELECT COUNT(*) FROM bacplus.bac WHERE an = {args.year}")
cnt = cur.fetchone()[0]

print(f"Deleting {cnt} entries from database...")

cur.execute("DELETE FROM bacplus.bac WHERE an = %s", (args.year,))

print(f"Inserting {len(data)} entries into database...")

chunk_size = 5000
chunks = [data[i : i + chunk_size] for i in range(0, len(data), chunk_size)]

for chunk in tqdm(chunks):
    cur.execute(
        f"INSERT INTO bacplus.bac({','.join(data_cols)}) VALUES "
        + ",".join(
            cur.mogrify(
                "(" + ",".join(["%s"] * len(x)) + ")",
                x,
            ).decode("utf-8")
            for x in chunk
        )
    )

conn.commit()
conn.close()
