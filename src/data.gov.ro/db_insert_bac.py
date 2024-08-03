import argparse
from openpyxl import load_workbook
import re
from dotenv import load_dotenv
import os
import psycopg2
import unidecode
from tqdm import tqdm

load_dotenv()


parser = argparse.ArgumentParser()
parser.add_argument("year", type=int)
args = parser.parse_args()

year = args.year
file = f"data/bac/data.gov.ro/{year}.xlsx"

# Load XLSX file

workbook = load_workbook(filename=file)
sheet = workbook.active

num_rows = sheet.max_row
num_cols = sheet.max_column

print("num_cols:", num_cols)

start_row = 1
while start_row <= num_rows and sheet.cell(row=start_row, column=2).value == None:
    start_row += 1

# Read data

header = [sheet.cell(row=start_row, column=i).value for i in range(1, num_cols + 1)]
column_names = [
    unidecode.unidecode(
        h.replace(".", "")
        .replace(" ", "_")
        .replace("(", "")
        .replace(")", "")
        .lower()
        .replace("fileira", "filiera")  # fix typo
    )
    for h in header
]

print(header)
print(column_names)

data = []
for i in range(start_row + 1, num_rows + 1):
    row = []
    for j in range(1, num_cols + 1):
        cell = sheet.cell(row=i, column=j).value
        if cell == "" or cell == "-" or str(cell) == "-2":
            cell = None
        if cell == "Da":
            cell = True
        if cell == "Nu":
            cell = False
        # dd/mm/yyyy -> yyyy-mm-dd
        if re.match(r"\d{2}/\d{2}/\d{4}", str(cell)):
            cell = re.sub(r"(\d{2})/(\d{2})/(\d{4})", r"\3-\2-\1", str(cell))
        row.append(cell)
    row.append(year)
    data.append(row)

# Insert data into database

if os.getenv("DATABASE_URL") is None:
    print("Make sure to specify DATABASE_URL in .env file")
    exit(1)

print(f"Connecting to database...")

conn = psycopg2.connect(os.getenv("DATABASE_URL"))
cur = conn.cursor()

cur.execute(f"SELECT COUNT(*) FROM bacplus.bac WHERE an = {year}")
cnt = cur.fetchone()[0]

print(f"Deleting {cnt} entries from database...")

cur.execute("DELETE FROM bacplus.bac WHERE an = %s", (year,))

print(f"Inserting {len(data)} entries into database...")

column_names.append("an")

chunk_size = 5000
chunks = [data[i : i + chunk_size] for i in range(0, len(data), chunk_size)]

for chunk in tqdm(chunks):
    cur.execute(
        f"INSERT INTO bacplus.bac({','.join(column_names)}) VALUES "
        + ",".join(
            cur.mogrify(
                "(" + ",".join(["%s"] * len(x)) + ")",
                x,
            ).decode("utf-8")
            for x in chunk
        )
    )

conn.commit()
conn.close()
